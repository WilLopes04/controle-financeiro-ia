from datetime import datetime
from dateutil.relativedelta import relativedelta
import uuid
import os
import psycopg2
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict
from dotenv import load_dotenv
from libsql_client import create_client_sync
from app.google_sheets import (
    atualizar_aba,
    atualizar_resumo_categoria,
    atualizar_resumo_cartao,
    SHEET_EMPRESA_ID,
    SHEET_PESSOAL_ID
)


load_dotenv()

TURSO_DATABASE_URL = (
    os.getenv("TURSO_DATABASE_URL")
    .replace("libsql://", "https://")
)

TURSO_AUTH_TOKEN = os.getenv("TURSO_AUTH_TOKEN")

db = create_client_sync(
    url=TURSO_DATABASE_URL,
    auth_token=TURSO_AUTH_TOKEN
)

# Caminho fixo da pasta do projeto: D:\ControleFinanceiro
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


DATABASE_URL = os.getenv("DATABASE_URL", "").strip()

ARQUIVO_EMPRESA = os.path.join(BASE_DIR, "Financeiro_Empresa.xlsx")
ARQUIVO_PESSOAL = os.path.join(BASE_DIR, "Financeiro_Pessoal.xlsx")




# =========================
# CRIAR TABELAS
# =========================


# =========================
# CALCULAR MÊS DA FATURA
# =========================

def calcular_mes_fatura(data_compra, nome_cartao):

    resultado = db.execute(
        "SELECT dia_fechamento FROM cartoes WHERE nome = ?",
        [nome_cartao]
    ).rows

    if not resultado:
        return data_compra

    dia_fechamento = resultado[0][0]

    if data_compra.day > dia_fechamento:
        return data_compra + relativedelta(months=1)

    return data_compra

# =========================
# REGISTRAR TRANSAÇÃO
# =========================

def atualizar_mes(nome_mes):
    gerar_planilha(nome_mes)

def registrar_transacao(
    tipo_conta,
    tipo_movimento,
    categoria,
    forma_pagamento,
    descricao,
    valor,
    data=None,
    parcela_atual=1,
    total_parcelas=1,
    id_compra=None
):

    if data is None:
        data = datetime.now().strftime("%Y-%m-%d")

    if id_compra is None:
        id_compra = str(uuid.uuid4())

    db.execute("""
        INSERT INTO transacoes (
            data,
            tipo_conta,
            tipo_movimento,
            categoria,
            forma_pagamento,
            descricao,
            valor,
            parcela_atual,
            total_parcelas,
            id_compra
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        data,
        tipo_conta,
        tipo_movimento,
        categoria,
        forma_pagamento,
        descricao,
        valor,
        parcela_atual,
        total_parcelas,
        id_compra
    ])

    mes = datetime.strptime(data, "%Y-%m-%d").month

    meses = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Março",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro"
   }

    atualizar_mes(meses[mes]) 

# =========================
# REGISTRAR PARCELADO
# =========================

def registrar_parcelado(
    tipo_conta,
    tipo_movimento,
    categoria,
    forma_pagamento,
    descricao,
    valor_total,
    total_parcelas
):
    valor_parcela = valor_total / total_parcelas
    data_base = datetime.now()

    resultado = db.execute("SELECT nome FROM cartoes")
    cartoes_validos = [linha[0] for linha in resultado.rows]

    if forma_pagamento in cartoes_validos:
        data_base = calcular_mes_fatura(data_base, forma_pagamento)

    id_compra = str(uuid.uuid4())

    for i in range(total_parcelas):
        nova_data = data_base + relativedelta(months=i)

        registrar_transacao(
            tipo_conta,
            tipo_movimento,
            categoria,
            forma_pagamento,
            f"{descricao} ({i+1}/{total_parcelas})",
            valor_parcela,
            nova_data.strftime("%Y-%m-%d"),
            i + 1,
            total_parcelas,
            id_compra
        )

# =========================
# GERAR PLANILHAS
# =========================



def gerar_planilha(mes_especifico=None):

    resultado = db.execute("SELECT * FROM transacoes")
    dados = resultado.rows

    meses = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março",
        4: "Abril", 5: "Maio", 6: "Junho",
        7: "Julho", 8: "Agosto", 9: "Setembro",
        10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }

    wb_empresa = Workbook()
    wb_pessoal = Workbook()
    wb_empresa.remove(wb_empresa.active)
    wb_pessoal.remove(wb_pessoal.active)

    for num_mes, nome_mes in meses.items():

        if mes_especifico and nome_mes != mes_especifico:
            continue

        dados_google_emp_linhas = []
        dados_google_pes_linhas = []	
        ws_emp = wb_empresa.create_sheet(nome_mes)
        ws_pes = wb_pessoal.create_sheet(nome_mes)


        cabecalho = [
            "Data", "Tipo Movimento", "Categoria",
            "Forma Pagamento", "Descrição",
            "Valor", "Parcela", "Total Parcelas"
        ]

        # ===== Cabeçalho Superior =====
        for ws in [ws_emp, ws_pes]:
            ws.append(["Total Entradas:", 0])
            ws.append(["Total Saídas:", 0])
            ws.append(["Saldo:", 0])
            ws.append([])
            ws.append(cabecalho)

        entradas_emp = 0
        saidas_emp = 0
        entradas_pes = 0
        saidas_pes = 0

        resumo_categoria_emp = defaultdict(float)
        resumo_categoria_pes = defaultdict(float)

        resumo_cartao_emp = defaultdict(float)
        resumo_cartao_pes = defaultdict(float)

        for linha in dados:
            data_transacao = datetime.strptime(linha[1], "%Y-%m-%d")

            if data_transacao.month == num_mes:

                linha_excel = [
                    linha[1], linha[3], linha[4],
                    linha[5], linha[6], linha[7],
                    linha[8], linha[9]
                ]

                tipo_conta = linha[2].lower()
                tipo_movimento = linha[3].lower()
                categoria = linha[4]
                forma_pagamento = linha[5]
                valor = float(linha[7])

                if tipo_conta == "empresa":
                    ws_emp.append(linha_excel)
                    dados_google_emp_linhas.append(linha_excel)

                    if tipo_movimento == "entrada":
                        entradas_emp += valor
                    else:
                        saidas_emp += valor
                        resumo_categoria_emp[categoria] += valor
                        resumo_cartao_emp[forma_pagamento] += valor

                else:
                    ws_pes.append(linha_excel)
                    dados_google_pes_linhas.append(linha_excel)

                    if tipo_movimento == "entrada":
                        entradas_pes += valor
                    else:
                        saidas_pes += valor
                        resumo_categoria_pes[categoria] += valor
                        resumo_cartao_pes[forma_pagamento] += valor

        # ===== Atualizar Totais =====
        ws_emp["B1"] = entradas_emp
        ws_emp["B2"] = saidas_emp
        ws_emp["B3"] = entradas_emp - saidas_emp

        ws_pes["B1"] = entradas_pes
        ws_pes["B2"] = saidas_pes
        ws_pes["B3"] = entradas_pes - saidas_pes

        dados_google_emp = [
            ["Total Entradas:", entradas_emp],
            ["Total Saídas:", saidas_emp],
            ["Saldo:", entradas_emp - saidas_emp],
            [],
            cabecalho
        ] + dados_google_emp_linhas

        dados_google_pes = [
            ["Total Entradas:", entradas_pes],
            ["Total Saídas:", saidas_pes],
            ["Saldo:", entradas_pes - saidas_pes],
            [],
            cabecalho
        ] + dados_google_pes_linhas

# ===== RESUMOS GOOGLE EMPRESA =====

        resumo_categoria_emp_google = [
            ["Resumo por Categoria"],
            ["Categoria", "Total"]
        ]

        for categoria, total in resumo_categoria_emp.items():
            resumo_categoria_emp_google.append(
            [categoria, total]
        )

        resumo_cartao_emp_google = [
            ["Resumo por Cartão"],
            ["Cartão", "Total Fatura"]
        ]

        for cartao, total in resumo_cartao_emp.items():
            resumo_cartao_emp_google.append(
            [cartao, total]
        )

# ===== RESUMOS GOOGLE PESSOAL =====

        resumo_categoria_pes_google = [
            ["Resumo por Categoria"],
            ["Categoria", "Total"]
        ]

        for categoria, total in resumo_categoria_pes.items():
            resumo_categoria_pes_google.append(
            [categoria, total]
        )

        resumo_cartao_pes_google = [
            ["Resumo por Cartão"],
            ["Cartão", "Total Fatura"]
        ]

        for cartao, total in resumo_cartao_pes.items():
            resumo_cartao_pes_google.append(
            [cartao, total]
        )
   

         # ===== Criar Tabela com Filtro =====
        for ws in [ws_emp, ws_pes]:
            if ws.max_row > 5:
                tabela = Table(
                    displayName=f"Tabela_{nome_mes}_{id(ws)}",
                    ref=f"A5:H{ws.max_row}"
                )
                estilo = TableStyleInfo(
                    name="TableStyleMedium9",
                    showRowStripes=True
                )
                tabela.tableStyleInfo = estilo
                ws.add_table(tabela)

        # ===== Resumo Lateral EMPRESA =====
        ws_emp["J2"] = "Resumo por Categoria"
        ws_emp["J3"] = "Categoria"
        ws_emp["K3"] = "Total"

        linha_res = 4
        for cat, total in resumo_categoria_emp.items():
            ws_emp[f"J{linha_res}"] = cat
            ws_emp[f"K{linha_res}"] = total
            linha_res += 1

        ws_emp["M2"] = "Resumo por Cartão"
        ws_emp["M3"] = "Cartão"
        ws_emp["N3"] = "Total Fatura"

        linha_res = 4
        for cartao, total in resumo_cartao_emp.items():
            ws_emp[f"M{linha_res}"] = cartao
            ws_emp[f"N{linha_res}"] = total
            linha_res += 1

        # ===== Resumo Lateral PESSOAL =====
        ws_pes["J2"] = "Resumo por Categoria"
        ws_pes["J3"] = "Categoria"
        ws_pes["K3"] = "Total"

        linha_res = 4
        for cat, total in resumo_categoria_pes.items():
            ws_pes[f"J{linha_res}"] = cat
            ws_pes[f"K{linha_res}"] = total
            linha_res += 1

        ws_pes["M2"] = "Resumo por Cartão"
        ws_pes["M3"] = "Cartão"
        ws_pes["N3"] = "Total Fatura"

        linha_res = 4
        for cartao, total in resumo_cartao_pes.items():
            ws_pes[f"M{linha_res}"] = cartao
            ws_pes[f"N{linha_res}"] = total
            linha_res += 1

        # ===== ENVIA PARA GOOGLE SHEETS =====

        atualizar_aba(
            SHEET_EMPRESA_ID,
            nome_mes.upper(),
            dados_google_emp
        )

        atualizar_aba(
            SHEET_PESSOAL_ID,
            nome_mes.upper(),
            dados_google_pes
        )

        atualizar_resumo_categoria(
            SHEET_EMPRESA_ID,
            nome_mes.upper(),
            resumo_categoria_emp_google
        )

        atualizar_resumo_cartao(
            SHEET_EMPRESA_ID,
            nome_mes.upper(),
            resumo_cartao_emp_google
        )

        atualizar_resumo_categoria(
            SHEET_PESSOAL_ID,
            nome_mes.upper(),
            resumo_categoria_pes_google
        )

        atualizar_resumo_cartao(
            SHEET_PESSOAL_ID,
            nome_mes.upper(),
            resumo_cartao_pes_google
        )

    wb_empresa.save(ARQUIVO_EMPRESA)
    wb_pessoal.save(ARQUIVO_PESSOAL)
# =========================
# CONSULTAS (NOVO)
# =========================

def total_entradas(tipo_conta):

    resultado = db.execute("""
        SELECT COALESCE(SUM(valor),0)
        FROM transacoes
        WHERE tipo_conta=? AND tipo_movimento='entrada'
    """, [tipo_conta]).rows

    return resultado[0][0]


def total_saidas(tipo_conta):

    resultado = db.execute("""
        SELECT COALESCE(SUM(valor),0)
        FROM transacoes
        WHERE tipo_conta=? AND tipo_movimento='saida'
    """, [tipo_conta]).rows

    return resultado[0][0]


def calcular_saldo(tipo_conta):
    return total_entradas(tipo_conta) - total_saidas(tipo_conta)


def fatura_cartao(nome_cartao, tipo_conta):

    resultado = db.execute("""
        SELECT COALESCE(SUM(valor),0)
        FROM transacoes
        WHERE tipo_conta=?
        AND tipo_movimento='saida'
        AND forma_pagamento=?
    """, [
        tipo_conta,
        nome_cartao
    ]).rows

    return resultado[0][0]

# =========================
# CALCULAR SALDO
# =========================

def calcular_saldo(tipo_conta):

    registros = db.execute("""
        SELECT tipo_movimento, valor
        FROM transacoes
        WHERE LOWER(tipo_conta) = LOWER(?)
    """, [tipo_conta]).rows

    saldo = 0

    for tipo_movimento, valor in registros:
        if tipo_movimento.lower() == "entrada":
            saldo += valor
        else:
            saldo -= valor

    return saldo

def listar_regras():

    rows = db.execute("""
        SELECT padrao, tipo_conta, categoria
        FROM regras
        ORDER BY padrao
    """).rows

    return rows

def salvar_regra(padrao: str, tipo_conta: str, categoria: str = None):

    padrao = (padrao or "").strip().lower()
    tipo_conta = (tipo_conta or "").strip().lower()
    categoria = (categoria or "").strip().lower() if categoria else None

    db.execute("""
        INSERT INTO regras (padrao, tipo_conta, categoria)
        VALUES (?, ?, ?)
        ON CONFLICT(padrao) DO UPDATE SET
            tipo_conta=excluded.tipo_conta,
            categoria=excluded.categoria
    """, [
        padrao,
        tipo_conta,
        categoria
    ])
def aplicar_regras(texto: str):
    t = (texto or "").lower()
    regras = listar_regras()

    for padrao, tipo_conta, categoria in regras:
        if padrao and padrao in t:
            out = {"tipo_conta": tipo_conta}
            if categoria:
                out["categoria"] = categoria
            return out

def registrar_transacao_turso(
    tipo_conta,
    tipo_movimento,
    categoria,
    forma_pagamento,
    descricao,
    valor,
    data=None,
    parcela_atual=1,
    total_parcelas=1,
    id_compra=None
):

    if data is None:
        data = datetime.now().strftime("%Y-%m-%d")

    if id_compra is None:
        id_compra = str(uuid.uuid4())

    db.execute("""
        INSERT INTO transacoes (
            data,
            tipo_conta,
            tipo_movimento,
            categoria,
            forma_pagamento,
            descricao,
            valor,
            parcela_atual,
            total_parcelas,
            id_compra
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        data,
        tipo_conta,
        tipo_movimento,
        categoria,
        forma_pagamento,
        descricao,
        valor,
        parcela_atual,
        total_parcelas,
        id_compra
    ])

    print("Transação gravada no Turso")

    return {}